"""
analyze.py  —  Skill Tracker (Project 2)
-----------------------------------------
Runs five SQL analyses against the skill_tracker database
and saves results to skill_analysis.xlsx.

Run: python3 analyze.py
"""

import os
import sys

import psycopg2
import pandas as pd
from tabulate import tabulate
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


DB_CONFIG = {
    "host":     "localhost",
    "port":     5432,
    "database": "skill_tracker",
    "user":      "postgres",
    "password": "",
}

OUTPUT_FILE = "skill_analysis.xlsx"


def get_conn():
    try:
        return psycopg2.connect(**DB_CONFIG)
    except psycopg2.OperationalError as e:
        print(f"connection error: {e}")
        sys.exit(1)


def run_query(conn, sql):
    return pd.read_sql_query(sql, conn)


# ── the five analyses ──────────────────────────────────────────────────────────

def top_skills_by_demand(conn):
    """most common skills across all postings, with avg salary where available"""
    return run_query(conn, """
        SELECT
            s.skill_name                                  AS skill,
            COUNT(*)                                      AS job_count,
            ROUND(AVG(j.salary_usd)::numeric, 0)          AS avg_salary_usd,
            COUNT(j.salary_usd)                           AS listings_with_salary
        FROM job_skills js
        JOIN skills s ON js.skill_id = s.id
        JOIN jobs   j ON js.job_id   = j.id
        GROUP BY s.skill_name
        ORDER BY job_count DESC
        LIMIT 15;
    """)


def skill_pairs_by_frequency(conn):
    """which two skills are listed together most often in the same job posting"""
    return run_query(conn, """
        SELECT
            s1.skill_name                                 AS skill_1,
            s2.skill_name                                 AS skill_2,
            COUNT(*)                                      AS jobs_with_both,
            ROUND(AVG(j.salary_usd)::numeric, 0)          AS avg_salary_usd
        FROM job_skills js1
        JOIN job_skills js2 ON js1.job_id = js2.job_id
                           AND js1.skill_id < js2.skill_id
        JOIN skills s1 ON js1.skill_id = s1.id
        JOIN skills s2 ON js2.skill_id = s2.id
        JOIN jobs   j  ON js1.job_id   = j.id
        GROUP BY s1.skill_name, s2.skill_name
        HAVING COUNT(*) >= 10
        ORDER BY jobs_with_both DESC
        LIMIT 20;
    """)


def skill_pairs_by_salary(conn):
    """same pairs but sorted by average salary — what actually pays more"""
    return run_query(conn, """
        SELECT
            s1.skill_name                                 AS skill_1,
            s2.skill_name                                 AS skill_2,
            COUNT(*)                                      AS jobs_with_both,
            ROUND(AVG(j.salary_usd)::numeric, 0)          AS avg_salary_usd
        FROM job_skills js1
        JOIN job_skills js2 ON js1.job_id = js2.job_id
                           AND js1.skill_id < js2.skill_id
        JOIN skills s1 ON js1.skill_id = s1.id
        JOIN skills s2 ON js2.skill_id = s2.id
        JOIN jobs   j  ON js1.job_id   = j.id
        WHERE j.salary_usd IS NOT NULL
        GROUP BY s1.skill_name, s2.skill_name
        HAVING COUNT(*) >= 8
        ORDER BY avg_salary_usd DESC
        LIMIT 15;
    """)


def city_top_skills(conn):
    """for each city, which 5 skills appear most in postings"""
    return run_query(conn, """
        WITH ranked AS (
            SELECT
                j.city,
                s.skill_name,
                COUNT(*)                                  AS job_count,
                ROUND(AVG(j.salary_usd)::numeric, 0)      AS avg_salary_usd,
                ROW_NUMBER() OVER (
                    PARTITION BY j.city ORDER BY COUNT(*) DESC
                ) AS rn
            FROM job_skills js
            JOIN jobs   j ON js.job_id   = j.id
            JOIN skills s ON js.skill_id = s.id
            GROUP BY j.city, s.skill_name
        )
        SELECT city, skill_name, job_count, avg_salary_usd
        FROM ranked
        WHERE rn <= 5
        ORDER BY city, rn;
    """)


def salary_by_experience(conn):
    """average and median salary by experience level — only rows that have salary data"""
    return run_query(conn, """
        SELECT
            experience_level,
            employment_type,
            COUNT(*)                                                  AS job_count,
            ROUND(AVG(salary_usd)::numeric, 0)                       AS avg_salary_usd,
            ROUND(
                PERCENTILE_CONT(0.5) WITHIN GROUP (ORDER BY salary_usd)::numeric, 0
            )                                                         AS median_salary_usd
        FROM jobs
        WHERE salary_usd IS NOT NULL
        GROUP BY experience_level, employment_type
        ORDER BY
            CASE experience_level
                WHEN 'Junior' THEN 1 WHEN 'Mid'    THEN 2
                WHEN 'Senior' THEN 3 WHEN 'Lead'   THEN 4
            END,
            employment_type;
    """)


# ── excel output ───────────────────────────────────────────────────────────────

def _bdr():
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name, df, hdr_hex, title):
    if name in wb.sheetnames:
        del wb[name]
    ws = wb.create_sheet(name)

    ncols = len(df.columns)
    ws.merge_cells(f"A1:{get_column_letter(ncols)}1")
    ws["A1"].value = title
    ws["A1"].font  = Font(bold=True, size=13, color=hdr_hex)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 5

    hfill = PatternFill("solid", fgColor=hdr_hex)
    afill = PatternFill("solid", fgColor="F5F8FF")

    for ci, col in enumerate(df.columns, 1):
        c = ws.cell(row=3, column=ci, value=col.replace("_", " ").title())
        c.fill = hfill
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()
    ws.row_dimensions[3].height = 20

    for ri, (_, row) in enumerate(df.iterrows(), start=4):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.border = _bdr()
            is_num = isinstance(val, (int, float))
            c.alignment = Alignment(
                horizontal="center" if (is_num or ci > 1) else "left",
                vertical="center"
            )
            if is_num and "salary" in df.columns[ci - 1].lower():
                c.number_format = "$#,##0"
            if ri % 2 == 0:
                c.fill = afill

    for ci, col in enumerate(df.columns, 1):
        max_len = max(len(str(col)), df.iloc[:, ci - 1].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 32)


def save_excel(results):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sheets = [
        ("Top Skills",           results["top_skills"],    "1A5276", "Top 15 In-Demand Skills (by job count)"),
        ("Skill Pairs Freq",     results["pairs_freq"],    "1F618D", "Top 20 Skill Combinations by Co-occurrence"),
        ("Skill Pairs Salary",   results["pairs_salary"],  "117A65", "Top 15 Skill Combos by Avg Salary (USD)"),
        ("City Rankings",        results["city_skills"],   "6E2F7D", "City-wise Top 5 Skills"),
        ("Experience vs Salary", results["exp_salary"],    "784212", "Salary by Experience Level (USD)"),
    ]
    for name, df, color, title in sheets:
        write_sheet(wb, name, df, color, title)

    wb.save(OUTPUT_FILE)
    print(f"\nresults saved  →  {OUTPUT_FILE}")


# ── main ───────────────────────────────────────────────────────────────────────

def section(title, df):
    print(f"\n{'─' * 58}")
    print(f"  {title}")
    print(f"{'─' * 58}")
    print(tabulate(df, headers="keys", tablefmt="rounded_outline", showindex=False))


def main():
    print("connecting to skill_tracker database...")
    conn = get_conn()

    results = {}
    print("running analysis 1/5: top skills by demand...")
    results["top_skills"]   = top_skills_by_demand(conn)

    print("running analysis 2/5: skill pairs by frequency...")
    results["pairs_freq"]   = skill_pairs_by_frequency(conn)

    print("running analysis 3/5: skill pairs by salary...")
    results["pairs_salary"] = skill_pairs_by_salary(conn)

    print("running analysis 4/5: city-wise top skills...")
    results["city_skills"]  = city_top_skills(conn)

    print("running analysis 5/5: salary by experience...")
    results["exp_salary"]   = salary_by_experience(conn)

    conn.close()

    section("Top 15 Skills by Demand",            results["top_skills"])
    section("Top 20 Skill Pairs by Co-occurrence", results["pairs_freq"])
    section("Top 15 Skill Pairs by Avg Salary",    results["pairs_salary"])
    section("City-wise Top 5 Skills",              results["city_skills"])
    section("Salary by Experience Level",          results["exp_salary"])

    save_excel(results)
    print("open skill_analysis.xlsx to see the formatted version.\n")


if __name__ == "__main__":
    main()
